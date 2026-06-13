from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from . import services
from .models import TimeEntry, VacationRequest


def export_user_summary_excel(
    *,
    rows: Iterable[dict[str, object]],
    totals: dict[str, int],
    period_range: str,
) -> BytesIO:
    """Per-user evaluation: one user per row, decimal hours for easy
    further processing (pivot tables, payroll, …)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Benutzerauswertung"

    bold = Font(bold=True)
    ws.append([f"Benutzerauswertung – Zeitraum {period_range}"])
    ws["A1"].font = bold
    ws.append([])

    headers = [
        "Benutzer",
        "Benutzername",
        "Buchungen",
        "Arbeitszeit (Std)",
        "Pausen (Std)",
        "Soll (Std)",
        "Urlaub (Std)",
        "Überstundenabbau (Std)",
        "Über-/Minusstunden (Std)",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{header_row + 1}"

    def _hours(minutes: object) -> float:
        return round(int(minutes or 0) / 60.0, 2)

    hour_columns = range(4, len(headers) + 1)
    for row in rows:
        row_user = row.get("user")
        ws.append(
            [
                str(getattr(row_user, "full_name", "")) or "",
                str(getattr(row_user, "username", "")) or "",
                int(row.get("count", 0)),
                _hours(row.get("work_minutes")),
                _hours(row.get("break_minutes")),
                _hours(row.get("target_minutes")),
                _hours(row.get("vacation_minutes")),
                _hours(row.get("overtime_taken_minutes")),
                _hours(row.get("balance_minutes")),
            ]
        )
        for column in hour_columns:
            ws.cell(row=ws.max_row, column=column).number_format = "0.00"

    ws.append(
        [
            "Summe",
            "",
            int(totals.get("count", 0)),
            _hours(totals.get("work_minutes")),
            _hours(totals.get("break_minutes")),
            _hours(totals.get("target_minutes")),
            _hours(totals.get("vacation_minutes")),
            _hours(totals.get("overtime_taken_minutes")),
            _hours(totals.get("balance_minutes")),
        ]
    )
    for cell in ws[ws.max_row]:
        cell.font = bold
    for column in hour_columns:
        ws.cell(row=ws.max_row, column=column).number_format = "0.00"

    for column_cells in ws.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None and cell.row > 1 else 0
            for cell in column_cells
        )
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_time_entries(
    entries: Iterable[TimeEntry],
    vacations: Optional[Iterable[VacationRequest]] = None,
    *,
    period_start: Optional[date] = None,
    period_end: Optional[date] = None,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Arbeitszeiten"

    headers = [
        "Mitarbeiter",
        "Firma",
        "Datum",
        "Start",
        "Ende",
        "Pause (Min)",
        "Arbeitszeit (Min)",
        "Überstunden (Min)",
        "Kommentar",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for entry in entries:
        end_value = "läuft" if entry.is_open else entry.end_time.strftime("%H:%M")
        ws.append(
            [
                entry.user.full_name if entry.user else "",
                entry.company.name if entry.company else "",
                entry.work_date.strftime("%d.%m.%Y"),
                entry.start_time.strftime("%H:%M"),
                end_value,
                entry.total_break_minutes,
                entry.worked_minutes,
                entry.overtime_minutes,
                entry.notes,
            ]
        )

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = max_length + 2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    vacation_list = list(vacations or [])
    if vacation_list:
        vacation_ws = wb.create_sheet(title="Urlaub")
        vacation_ws.append(["Start", "Ende", "Anzurechnung (Min)", "Typ", "Kommentar"])
        for cell in vacation_ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for vacation in vacation_list:
            start_date = max(period_start or vacation.start_date, vacation.start_date)
            end_date = min(period_end or vacation.end_date, vacation.end_date)
            credited = services.calculate_required_vacation_minutes(
                vacation.user,
                start_date,
                end_date,
            )
            if credited <= 0:
                continue
            vacation_ws.append(
                [
                    start_date.strftime("%d.%m.%Y"),
                    end_date.strftime("%d.%m.%Y"),
                    credited,
                    "Überstundenabbau" if vacation.use_overtime else "Urlaub",
                    vacation.comment or "",
                ]
            )
        for column_cells in vacation_ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = max_length + 2
            vacation_ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
