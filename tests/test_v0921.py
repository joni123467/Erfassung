"""Regression tests for 0.9.21 – Einsatzort einer Buchung (Remote / vor Ort).

Covers: version bump, both new columns plus migration 13, the per-user
switch (``users.remote_flag_enabled``) gating the field, clocking in remote via
``/punch`` (and correcting it afterwards through the comment step), manual
bookings via ``/time``, the admin edit form, the flag surviving entry splits,
the mobile sync payload, and the conditional „Ort" column in the PDF/Excel
exports.
"""

from __future__ import annotations

import re
import sys
from datetime import date, time

import pytest

import licensed_env


def _fresh_app(tmp_path, monkeypatch, env: dict | None = None):
    monkeypatch.setenv("ERFASSUNG_CONFIG_DIR", str(tmp_path / "config"))
    monkeypatch.setenv("ERFASSUNG_LOGS_DIR", str(tmp_path / "logs"))
    monkeypatch.setenv("ERFASSUNG_DATA_DIR", str(tmp_path / "data"))
    monkeypatch.setenv("SESSION_SECRET_KEY", "test-secret")
    monkeypatch.setenv("ERFASSUNG_DISABLE_SCHEDULER", "1")
    for key in ("DATABASE_URL", "DB_TYPE", "DB_HOST", "DB_PORT", "DB_NAME",
                "DB_USER", "DB_PASSWORD", "DB_SSL", "DB_PATH"):
        monkeypatch.delenv(key, raising=False)
    for key, value in (env or {}).items():
        monkeypatch.setenv(key, value)
    for name in [m for m in sys.modules if m.startswith("app")]:
        del sys.modules[name]
    import app.main as main

    licensed_env.activate()
    return main


@pytest.fixture()
def client(tmp_path, monkeypatch):
    main = _fresh_app(tmp_path, monkeypatch, {"DATABASE_URL": f"sqlite:///{tmp_path}/erfassung.db"})
    from fastapi.testclient import TestClient

    with TestClient(main.app) as test_client:
        from app import crud, database, security

        db = database.SessionLocal()
        try:
            admin = crud.get_user_by_username(db, "admin")
            admin.password_hash = security.hash_password("Admin!0000")
            admin.must_change_password = False
            db.commit()
        finally:
            db.close()
        test_client.main = main  # type: ignore[attr-defined]
        yield test_client


_CSRF_RE = re.compile(r'name="csrf_token" value="([^"]+)"')


def _csrf(client, url: str) -> str:
    html = client.get(url).text
    m = _CSRF_RE.search(html)
    assert m, f"no csrf token on {url}"
    return m.group(1)


def login(client, username: str = "admin", password: str = "Admin!0000"):
    token = _csrf(client, "/login")
    return client.post(
        "/login",
        data={"username": username, "password": password, "csrf_token": token},
        follow_redirects=False,
    )


DAY = date(2026, 8, 12)


def _admin_id() -> int:
    from app import crud, database

    db = database.SessionLocal()
    try:
        return crud.get_user_by_username(db, "admin").id
    finally:
        db.close()


def _enable_remote(enabled: bool = True) -> None:
    from app import crud, database

    db = database.SessionLocal()
    try:
        crud.get_user_by_username(db, "admin").remote_flag_enabled = enabled
        db.commit()
    finally:
        db.close()


def _entries(user_id):
    from app import database, models

    db = database.SessionLocal()
    try:
        return (
            db.query(models.TimeEntry)
            .filter(models.TimeEntry.user_id == user_id)
            .order_by(models.TimeEntry.start_time)
            .all()
        )
    finally:
        db.close()


def _entry(entry_id):
    from app import crud, database

    db = database.SessionLocal()
    try:
        return crud.get_time_entry(db, entry_id)
    finally:
        db.close()


def _closed_entry(user_id, start, end, *, is_remote=False, notes="Büro", day=DAY):
    from app import crud, database, models, schemas

    db = database.SessionLocal()
    try:
        return crud.create_time_entry(
            db,
            schemas.TimeEntryCreate(
                user_id=user_id, company_id=None, work_date=day,
                start_time=start, end_time=end, break_minutes=0,
                break_started_at=None, is_open=False, notes=notes,
                status=models.TimeEntryStatus.APPROVED, is_manual=False,
                is_remote=is_remote,
            ),
        ).id
    finally:
        db.close()


# --- version & schema ------------------------------------------------------------

def test_version(client):
    assert client.main.APP_VERSION == "0.17.0"
    assert client.get("/health").json()["version"] == "0.17.0"


def test_columns_exist(client):
    from sqlalchemy import inspect

    from app import database

    inspector = inspect(database.engine)
    users = {c["name"] for c in inspector.get_columns("users")}
    entries = {c["name"] for c in inspector.get_columns("time_entries")}
    assert "remote_flag_enabled" in users
    assert "is_remote" in entries


def test_migration_registered(client):
    from app import db_migrations

    versions = [version for version, _ in db_migrations.MIGRATIONS]
    assert 13 in versions
    assert versions == sorted(versions)


def test_default_is_on_site(client):
    """Ohne Angabe gilt eine Buchung als vor Ort."""
    entry = _entry(_closed_entry(_admin_id(), time(8, 0), time(12, 0)))
    assert entry.is_remote is False
    assert entry.location_label == "Vor Ort"


# --- clocking --------------------------------------------------------------------

def test_punch_start_work_remote(client):
    _enable_remote()
    login(client)
    token = _csrf(client, "/dashboard")
    response = client.post(
        "/punch",
        data={"csrf_token": token, "action": "start_work", "is_remote": "1",
              "next_url": "/dashboard"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    entries = _entries(_admin_id())
    assert len(entries) == 1
    assert entries[0].is_remote is True
    assert entries[0].location_label == "Remote"


def test_punch_start_work_without_flag_is_on_site(client):
    _enable_remote()
    login(client)
    token = _csrf(client, "/dashboard")
    client.post(
        "/punch",
        data={"csrf_token": token, "action": "start_work", "next_url": "/dashboard"},
        follow_redirects=False,
    )
    assert _entries(_admin_id())[0].is_remote is False


def test_punch_remote_ignored_when_not_enabled(client):
    """Ohne Freischaltung wird ein mitgesendeter Haken nicht übernommen."""
    login(client)
    token = _csrf(client, "/dashboard")
    client.post(
        "/punch",
        data={"csrf_token": token, "action": "start_work", "is_remote": "1",
              "next_url": "/dashboard"},
        follow_redirects=False,
    )
    assert _entries(_admin_id())[0].is_remote is False


def test_punch_update_notes_corrects_location(client):
    """Der Kommentar-Nachtrag korrigiert auch den Einsatzort."""
    _enable_remote()
    login(client)
    uid = _admin_id()
    entry_id = _closed_entry(uid, time(8, 0), time(12, 0), is_remote=False)
    token = _csrf(client, "/dashboard")
    client.post(
        "/punch",
        data={"csrf_token": token, "action": "update_notes", "entry_id": str(entry_id),
              "notes": "Telefonat", "is_remote": "1", "next_url": "/dashboard"},
        follow_redirects=False,
    )
    updated = _entry(entry_id)
    assert updated.is_remote is True and updated.notes == "Telefonat"


def test_update_notes_keeps_location_when_not_enabled(client):
    """Ohne Freischaltung bleibt ein bereits gesetzter Einsatzort erhalten."""
    login(client)
    uid = _admin_id()
    entry_id = _closed_entry(uid, time(8, 0), time(12, 0), is_remote=True)
    token = _csrf(client, "/dashboard")
    client.post(
        "/punch",
        data={"csrf_token": token, "action": "update_notes", "entry_id": str(entry_id),
              "notes": "Neu", "next_url": "/dashboard"},
        follow_redirects=False,
    )
    assert _entry(entry_id).is_remote is True


def test_manual_entry_remote(client):
    _enable_remote()
    login(client)
    token = _csrf(client, "/dashboard")
    response = client.post(
        "/time",
        data={"csrf_token": token, "work_date": DAY.isoformat(), "start_time": "09:00",
              "end_time": "09:30", "break_minutes": "0", "notes": "Telefonat",
              "is_remote": "1", "next_url": "/dashboard"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    entries = _entries(_admin_id())
    assert len(entries) == 1 and entries[0].is_remote is True


# --- splits keep the flag --------------------------------------------------------

def test_split_of_closed_entry_keeps_location(client):
    from app import crud, database, models, schemas

    uid = _admin_id()
    _closed_entry(uid, time(8, 0), time(12, 0), is_remote=True)
    db = database.SessionLocal()
    try:
        crud.create_manual_time_entry(
            db,
            schemas.TimeEntryCreate(
                user_id=uid, company_id=None, work_date=DAY,
                start_time=time(10, 0), end_time=time(10, 20), break_minutes=0,
                break_started_at=None, is_open=False, notes="Telefonat",
                status=models.TimeEntryStatus.PENDING, is_manual=True, is_remote=True,
            ),
        )
    finally:
        db.close()
    entries = _entries(uid)
    assert len(entries) == 3
    assert all(entry.is_remote for entry in entries)


def test_split_of_running_entry_keeps_location(client):
    """Eine laufende Buchung reicht bis „jetzt"; der Nachtrag liegt deshalb
    bewusst am Tagesanfang, damit der Test unabhängig von der Uhrzeit läuft."""
    from datetime import datetime

    from app import crud, database, models, schemas

    today = date.today()
    uid = _admin_id()
    db = database.SessionLocal()
    try:
        crud.start_running_entry(
            db, user_id=uid, started_at=datetime.combine(today, time(0, 0)), is_remote=True
        )
        crud.create_manual_time_entry(
            db,
            schemas.TimeEntryCreate(
                user_id=uid, company_id=None, work_date=today,
                start_time=time(0, 1), end_time=time(0, 2), break_minutes=0,
                break_started_at=None, is_open=False, notes="Telefonat",
                status=models.TimeEntryStatus.PENDING, is_manual=True, is_remote=True,
            ),
        )
    finally:
        db.close()
    entries = _entries(uid)
    assert len(entries) == 3
    # Der abgeschlossene erste Abschnitt übernimmt den Einsatzort der laufenden Buchung
    assert all(entry.is_remote for entry in entries)


# --- admin edit ------------------------------------------------------------------

def test_admin_form_shows_checkbox_only_when_enabled(client):
    login(client)
    uid = _admin_id()
    entry_id = _closed_entry(uid, time(8, 0), time(12, 0))
    url = f"/admin/time-entries/{entry_id}/edit?next=/admin/reports/time&user={uid}"

    assert 'name="is_remote"' not in client.get(url).text
    _enable_remote()
    assert 'name="is_remote"' in client.get(url).text


def test_admin_update_sets_and_clears_location(client):
    _enable_remote()
    login(client)
    uid = _admin_id()
    entry_id = _closed_entry(uid, time(8, 0), time(12, 0))
    url = f"/admin/time-entries/{entry_id}/edit?next=/admin/reports/time&user={uid}"
    base = {
        "user_id": str(uid), "work_date": DAY.isoformat(), "start_time": "08:00",
        "end_time": "12:00", "break_minutes": "0", "notes": "Büro",
        "next_url": "/admin/reports/time",
    }

    client.post(f"/admin/time-entries/{entry_id}/update",
                data={**base, "csrf_token": _csrf(client, url), "is_remote": "1",
                      "change_reason": "Test: Korrektur"},
                follow_redirects=False)
    assert _entry(entry_id).is_remote is True

    client.post(f"/admin/time-entries/{entry_id}/update",
                data={**base, "csrf_token": _csrf(client, url),
                      "change_reason": "Test: Korrektur"},
                follow_redirects=False)
    assert _entry(entry_id).is_remote is False


def test_admin_update_keeps_location_when_not_enabled(client):
    """Ist das Feld für den Benutzer aus, bleibt der Bestandswert erhalten."""
    login(client)
    uid = _admin_id()
    entry_id = _closed_entry(uid, time(8, 0), time(12, 0), is_remote=False)
    url = f"/admin/time-entries/{entry_id}/edit?next=/admin/reports/time&user={uid}"
    client.post(
        f"/admin/time-entries/{entry_id}/update",
        data={"csrf_token": _csrf(client, url), "user_id": str(uid),
              "change_reason": "Test: Korrektur",
              "work_date": DAY.isoformat(), "start_time": "08:00", "end_time": "11:00",
              "break_minutes": "0", "notes": "Büro", "is_remote": "1",
              "next_url": "/admin/reports/time"},
        follow_redirects=False,
    )
    updated = _entry(entry_id)
    assert updated.is_remote is False and updated.end_time == time(11, 0)


def test_user_form_offers_setting(client):
    login(client)
    html = client.get("/admin/users/new").text
    assert 'name="remote_flag_enabled"' in html
    assert "Einsatzort erfassen" in html


# --- surfaces --------------------------------------------------------------------

def test_dashboard_checkbox_follows_setting(client):
    login(client)
    assert 'name="is_remote"' not in client.get("/dashboard").text
    _enable_remote()
    assert 'name="is_remote"' in client.get("/dashboard").text


def test_mobile_sync_exposes_location(client):
    _enable_remote()
    login(client)
    _closed_entry(_admin_id(), time(8, 0), time(12, 0), is_remote=True, day=date.today())
    payload = client.get("/mobile/sync-data").json()
    assert payload["permissions"]["flag_remote"] is True
    assert any(entry["is_remote"] for entry in payload["entries"])


def test_booking_table_shows_badge(client):
    login(client)
    _closed_entry(_admin_id(), time(8, 0), time(12, 0), is_remote=True)
    html = client.get(f"/records?month={DAY.strftime('%Y-%m')}").text
    assert "location-badge" in html and "Remote" in html


# --- exports ---------------------------------------------------------------------

def test_pdf_location_column_only_when_used(client):
    from app import crud, database, models
    from app.pdf_export import any_remote

    uid = _admin_id()
    _closed_entry(uid, time(8, 0), time(12, 0))
    db = database.SessionLocal()
    try:
        entries = crud.get_time_entries(db, uid, statuses=[models.TimeEntryStatus.APPROVED])
        assert any_remote(entries) is False
    finally:
        db.close()

    _closed_entry(uid, time(13, 0), time(15, 0), is_remote=True)
    db = database.SessionLocal()
    try:
        entries = crud.get_time_entries(db, uid, statuses=[models.TimeEntryStatus.APPROVED])
        assert any_remote(entries) is True
    finally:
        db.close()


def test_records_pdf_with_remote_entry(client):
    login(client)
    _closed_entry(_admin_id(), time(8, 0), time(12, 0), is_remote=True)
    response = client.get(f"/records/pdf?month={DAY.strftime('%Y-%m')}")
    assert response.status_code == 200
    assert response.content.startswith(b"%PDF")


def test_excel_location_column_only_when_used(client):
    from openpyxl import load_workbook

    from app import crud, database, models
    from app.excel_export import export_time_entries

    uid = _admin_id()
    _closed_entry(uid, time(8, 0), time(12, 0))
    db = database.SessionLocal()
    try:
        plain = export_time_entries(
            crud.get_time_entries(db, uid, statuses=[models.TimeEntryStatus.APPROVED])
        )
    finally:
        db.close()
    header = [cell.value for cell in load_workbook(plain).active[1]]
    assert "Ort" not in header

    _closed_entry(uid, time(13, 0), time(15, 0), is_remote=True)
    db = database.SessionLocal()
    try:
        detailed = export_time_entries(
            crud.get_time_entries(db, uid, statuses=[models.TimeEntryStatus.APPROVED])
        )
    finally:
        db.close()
    sheet = load_workbook(detailed).active
    header = [cell.value for cell in sheet[1]]
    assert header[2] == "Ort"
    assert {row[2].value for row in sheet.iter_rows(min_row=2)} == {"Vor Ort", "Remote"}
